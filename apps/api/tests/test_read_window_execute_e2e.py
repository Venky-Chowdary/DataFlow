"""The declared read window is what actually lands in a real destination.

Unit parity (preview == COUNT == iterator) proves the readers agree with each
other; it does not prove the *writer* moved that population. These tests run the
real engine into live PostgreSQL and then re-read the destination with a second
driver, so the rows asserted here are rows the database really holds.

Two properties, both client-visible:

* a workbook whose header sits on row 3 of the third sheet, with a totals
  footer, lands as its data rows only — no title row, no ``TOTAL`` row, no
  phantom sheet;
* a request that has been through the durable dict (which is what a scheduled
  beat and a restarted worker reconstruct) reads the same window, because a
  schedule that quietly reverts to "active sheet, row 1" would silently load a
  different population every night.
"""

from __future__ import annotations

import io
import uuid

import psycopg2
import pytest
from openpyxl import Workbook

from services.read_options import ReadOptions
from src.transfer.engine import UniversalTransferEngine
from src.transfer.models import (
    EndpointConfig,
    TransferRequest,
    transfer_request_from_dict,
    transfer_request_to_dict,
)
from tests.typed_fidelity_helpers import pg_endpoint, require_ports, uniq

pytestmark = pytest.mark.timeout(300)

# Title row, blank row, header on row 3, three data rows, totals footer.
MESSY_ROWS = [
    ["Quarterly export"],
    [],
    ["id", "name", "amount"],
    [1, "a", "10.50"],
    [2, "b", "20.50"],
    [3, "c", "30.50"],
    ["TOTAL", "", "61.50"],
]

MESSY_CSV = (
    b"Quarterly export\n"
    b"\n"
    b"id;name;amount\n"
    b"1;a;10.50\n"
    b"2;b;20.50\n"
    b"3;c;30.50\n"
    b"TOTAL;;61.50\n"
)


def _messy_workbook() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(title="Cover").append(["do not read me"])
    data = wb.create_sheet(title="Data")
    for row in MESSY_ROWS:
        data.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mappings() -> list[dict[str, object]]:
    return [
        {
            "source": name,
            "target": name,
            "target_type": target_type,
            "approved": True,
            "confidence": 0.99,
        }
        for name, target_type in (
            ("id", "BIGINT"),
            ("name", "TEXT"),
            # The profiler reads 10.50 as DECIMAL(7,4); a wider or textual
            # target is a fidelity collapse and G-gates refuse it, correctly.
            ("amount", "DECIMAL(7,4)"),
        )
    ]


_FORMATS = {"xlsx": "excel", "csv": "csv"}


def _request(
    content: bytes, filename: str, table: str, options: ReadOptions
) -> TransferRequest:
    return TransferRequest(
        source=EndpointConfig(
            kind="file", format=_FORMATS[filename.rsplit(".", 1)[-1]]
        ),
        destination=pg_endpoint(table),
        source_content=content,
        source_filename=filename,
        mappings=_mappings(),
        sync_mode="full_refresh_overwrite",
        validation_mode="strict",
        read_options=options.to_wire(),
    )


def _landed(table: str) -> list[tuple]:
    conn = psycopg2.connect(
        host="localhost",
        port=5432,
        database="dataflow",
        user="dataflow",
        password="dataflow",
    )
    try:
        with conn.cursor() as cur:
            cur.execute(f'SELECT id, name, amount FROM public."{table}" ORDER BY id')
            return [tuple(r) for r in cur.fetchall()]
    finally:
        conn.close()


def _drop(table: str) -> None:
    conn = psycopg2.connect(
        host="localhost",
        port=5432,
        database="dataflow",
        user="dataflow",
        password="dataflow",
    )
    try:
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute(f'DROP TABLE IF EXISTS public."{table}"')
    finally:
        conn.close()


def _execute(request: TransferRequest):
    return UniversalTransferEngine().execute_tracked(request, uuid.uuid4().hex[:24])


@pytest.fixture()
def pg_table():
    require_ports(5432, 27017)
    table = uniq("read_window")
    yield table
    _drop(table)


def test_excel_window_is_the_population_that_lands_in_postgres(pg_table):
    """Sheet 2, header row 3, footer excluded — and nothing else moves."""
    options = ReadOptions(sheet="Data", header_row=3, skip_footer=1)
    result = _execute(_request(_messy_workbook(), "quarterly.xlsx", pg_table, options))
    assert result.success, result.error
    assert result.records_transferred == 3, result.error
    landed = _landed(pg_table)
    assert [r[0] for r in landed] == [1, 2, 3]
    assert [r[1] for r in landed] == ["a", "b", "c"]
    assert [float(r[2]) for r in landed] == [10.50, 20.50, 30.50]


def test_csv_window_is_the_population_that_lands_in_postgres(pg_table):
    options = ReadOptions(header_row=3, skip_rows=1, skip_footer=1, delimiter=";")
    result = _execute(_request(MESSY_CSV, "quarterly.csv", pg_table, options))
    assert result.success, result.error
    assert result.records_transferred == 2, result.error
    assert [r[0] for r in _landed(pg_table)] == [2, 3]


def test_a_rebuilt_request_reads_the_same_window(pg_table, tmp_path):
    """What a scheduled beat and a restarted worker reconstruct still narrows.

    Durable serialization deliberately drops in-memory bytes and keeps the
    staged path, so the round trip here stages the workbook exactly as the
    upload path does before a schedule stores the request.
    """
    options = ReadOptions(sheet="Data", header_row=3, skip_footer=1)
    staged = tmp_path / "quarterly.xlsx"
    staged.write_bytes(_messy_workbook())
    original = _request(b"", "quarterly.xlsx", pg_table, options)
    original.source_path = str(staged)
    revived = transfer_request_from_dict(transfer_request_to_dict(original))
    assert revived.read_options == options.to_wire()
    assert revived.source_path == str(staged)
    result = _execute(revived)
    assert result.success, result.error
    assert result.records_transferred == 3, result.error
    assert [r[0] for r in _landed(pg_table)] == [1, 2, 3]


def test_an_unknown_sheet_refuses_before_anything_is_written(pg_table):
    """A typo must not fall back to the active sheet and load the cover page."""
    options = ReadOptions(sheet="Dta", header_row=3)
    result = _execute(_request(_messy_workbook(), "quarterly.xlsx", pg_table, options))
    assert result.success is False
    assert "Dta" in (result.error or "")
    with pytest.raises(psycopg2.Error):
        _landed(pg_table)
