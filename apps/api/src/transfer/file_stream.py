"""Streaming file → database transfer for CSV, TSV, JSONL, NDJSON, JSON arrays,
Excel, and Parquet.  Supports in-memory ``bytes`` as well as on-disk paths so
billion-row files can be processed without loading the whole payload into RAM.
"""

from __future__ import annotations

import csv
import gzip
import io
import itertools
import json
import logging
import os
from services.brand_env import getenv_brand
import sys
import tempfile
from collections.abc import Callable
from contextlib import contextmanager
from functools import partial
from pathlib import Path
from typing import Any

from .models import EndpointConfig
from .type_mapper import ddl_carrier_type, ddl_type

try:
    from services.checkpoint_service import Checkpoint, CheckpointService
    from services.error_handling import RetryBudget, with_retry
    from services.parallel_chunks import OrderedChunkRunner
    from services.replay_safety import classify_replay_safety
    from services.resilience import adaptive_chunk_size
    from services.row_filter import apply_row_filter
    from services.transform_engine import infer_date_locale, set_active_date_locale
except ImportError:  # pragma: no cover - tests with api root on path
    from src.services.checkpoint_service import Checkpoint, CheckpointService
    from src.services.error_handling import RetryBudget, with_retry
    from src.services.parallel_chunks import OrderedChunkRunner
    from src.services.replay_safety import classify_replay_safety
    from src.services.resilience import adaptive_chunk_size
    from src.services.row_filter import apply_row_filter
    from src.services.transform_engine import infer_date_locale, set_active_date_locale

_api_root = Path(__file__).resolve().parents[2]
if str(_api_root) not in sys.path:
    sys.path.insert(0, str(_api_root))

from connectors.writer_common import (
    CHUNK_SIZE,
    map_rows_for_fingerprint,
    resolve_target_columns,
    row_fingerprints,
    transform_error_policy_for_validation_mode,
)
from services.dest_precount import (
    PRECOUNT_KEY,
    begin_overwrite_source_keys,
    records_to_key_tuples,
    precount_table,
    stamp_overwrite_source_keys,
)
from services.excel_parser import sheet_headers
from services.reconciliation import FingerprintAccumulator
from services.tabular_rows import is_blank_row

try:
    from services.csv_profiler import (
        count_csv_rows,
        detect_delimiter,
        detect_encoding,
        parse_csv_preview,
    )
except ImportError:  # pragma: no cover - compatibility for tests with api root on PYTHONPATH
    from src.services.csv_profiler import (
        count_csv_rows,
        detect_delimiter,
        detect_encoding,
        parse_csv_preview,
    )

from .adapters import (
    WriteBatchBlocked,
    records_to_matrix,
    resolve_connector_config,
    resolve_dest_table,
)
from .stream import _write_batch

STREAMABLE_TYPES = {"csv", "tsv", "jsonl", "ndjson", "json", "excel", "parquet", "avro", "orc"}
STREAM_THRESHOLD = int(getenv_brand("STREAM_FILE_ROWS", "1"))
FILE_SPILL_THRESHOLD = int(getenv_brand("FILE_SPILL_THRESHOLD", str(50 * 1024 * 1024)))
SPILL_DIR = getenv_brand("SPILL_DIR") or None


_JSONL_SCALAR_ERROR = (
    "JSONL record must be a JSON object; scalar records require an explicit "
    "normalization step to avoid silent data loss."
)


def _is_path(value: Any) -> bool:
    return isinstance(value, (str, os.PathLike))


def _source_suffix(filename: str) -> str:
    name = os.path.basename(filename or "upload")
    _, ext = os.path.splitext(name)
    return ext or ".tmp"


def prepare_stream_content(
    content: bytes = b"",
    filename: str = "upload.csv",
    source_path: str = "",
) -> bytes | str:
    """Return the most efficient source reference for streaming.

    If an explicit ``source_path`` is provided and exists, it is used.
    If ``content`` is larger than ``FILE_SPILL_THRESHOLD`` bytes, it is written
    to a temporary file and that path is returned so iteration can stream from
    disk.  Otherwise the original ``bytes`` payload is returned.
    """
    if source_path and os.path.isfile(source_path):
        return source_path
    if not content:
        return content
    if len(content) <= FILE_SPILL_THRESHOLD:
        return content

    suffix = _source_suffix(filename)
    fd, path = tempfile.mkstemp(suffix=suffix, prefix="dataflow_spill_", dir=SPILL_DIR)
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(content)
    except Exception:
        os.close(fd)
        raise
    return path


def _is_gzip_bytes(sample: bytes) -> bool:
    return bool(sample) and sample[:2] == b"\x1f\x8b"


def _is_gzip_path(path: str | os.PathLike) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(2) == b"\x1f\x8b"
    except Exception:
        return False


def _first_bytes(content: bytes | str | os.PathLike, size: int = 8192) -> bytes:
    """Return a decompressed prefix for sniffing."""
    if _is_path(content):
        if _is_gzip_path(content):
            with gzip.open(content, "rb") as f:
                return f.read(size)
        with open(content, "rb") as f:
            return f.read(size)
    if isinstance(content, (bytes, bytearray)) and _is_gzip_bytes(content):
        with gzip.GzipFile(fileobj=io.BytesIO(content)) as f:
            return f.read(size)
    return bytes(content[:size])


def _open_binary(content: bytes | str | os.PathLike) -> Any:
    """Open a binary stream, transparently decompressing gzip payloads."""
    if _is_path(content):
        if _is_gzip_path(content):
            return gzip.open(content, "rb")
        return open(content, "rb")
    if isinstance(content, (bytes, bytearray)) and _is_gzip_bytes(content):
        return gzip.GzipFile(fileobj=io.BytesIO(content))
    return io.BytesIO(content)


@contextmanager
def _text_reader(content: bytes | str | os.PathLike, encoding: str | None = None, newline: str = ""):
    binary = _open_binary(content)
    text = None
    try:
        if encoding is None:
            encoding = detect_encoding(_first_bytes(content))
        text = io.TextIOWrapper(binary, encoding=encoding, errors="strict", newline=newline)
        yield text
    finally:
        if text is not None:
            try:
                text.close()
            except Exception as exc:
                logging.getLogger(__name__).debug("Exception suppressed: %s", exc, exc_info=exc)
        else:
            binary.close()


def _excel_preview(content: bytes | str | os.PathLike, preview_rows: int = 100) -> tuple[list[str], list[list[str]], int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Excel import is not ready on this platform node. Datawrap bundles file parsers — retry shortly."
        ) from exc

    wb = load_workbook(content, read_only=True, data_only=True) if _is_path(content) else load_workbook(
        io.BytesIO(content), read_only=True, data_only=True
    )
    try:
        ws = wb.active
        if ws is None:
            return [], [], 0

        row_iter = ws.iter_rows(values_only=True)
        first = next(row_iter, None)
        if not first:
            return [], [], 0

        headers = sheet_headers(first)
        preview: list[list[str]] = []
        total = 0

        for row in row_iter:
            if is_blank_row(row):
                continue
            total += 1
            if len(preview) < preview_rows:
                preview.append([str(c).strip() if c is not None else "" for c in row])

        return headers, preview, total
    finally:
        wb.close()


def _excel_batches(content: bytes | str | os.PathLike, chunk_size: int):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Excel import is not ready on this platform node. Datawrap bundles file parsers — retry shortly."
        ) from exc

    wb = load_workbook(content, read_only=True, data_only=True) if _is_path(content) else load_workbook(
        io.BytesIO(content), read_only=True, data_only=True
    )
    try:
        ws = wb.active
        if ws is None:
            return

        row_iter = ws.iter_rows(values_only=True)
        first = next(row_iter, None)
        if not first:
            return

        headers = sheet_headers(first)
        batch: list[dict] = []
        for row in row_iter:
            # A formatting-only row is not a record; writing it would land an
            # all-NULL row the source never had.
            if is_blank_row(row):
                continue
            # Wider data rows than the header silently lost trailing cells —
            # refuse rather than slice away columns (JSONL-style honesty).
            if len(row) > len(headers):
                raise ValueError(
                    f"Excel row has {len(row)} cells but header has {len(headers)} "
                    "columns; refuse silent column drop — widen the header row "
                    "or fix the sheet"
                )
            record = {
                headers[i]: ("" if c is None else str(c).strip())
                for i, c in enumerate(row[: len(headers)])
            }
            batch.append(record)
            if len(batch) >= chunk_size:
                yield batch
                batch = []
        if batch:
            yield batch
    finally:
        wb.close()


def _excel_count(content: bytes | str | os.PathLike) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Excel import is not ready on this platform node. Datawrap bundles file parsers — retry shortly."
        ) from exc

    wb = load_workbook(content, read_only=True, data_only=True) if _is_path(content) else load_workbook(
        io.BytesIO(content), read_only=True, data_only=True
    )
    try:
        ws = wb.active
        if ws is None:
            return 0
        row_iter = ws.iter_rows(values_only=True)
        if next(row_iter, None) is None:
            return 0
        # ``max_row`` is the used range, which formatting inflates.
        return sum(1 for row in row_iter if not is_blank_row(row))
    finally:
        wb.close()


def supports_file_streaming(source_kind: str, filename: str, destination: EndpointConfig) -> bool:
    if source_kind != "file" or destination.kind != "database":
        return False
    try:
        from services.file_parser import FileParser
    except ImportError:  # pragma: no cover - compatibility for tests with api root on PYTHONPATH
        from src.services.file_parser import FileParser

    return FileParser.detect_file_type(filename) in STREAMABLE_TYPES


def _decompress_bytes_if_gzip(data: bytes) -> bytes:
    """Decompress an in-memory gzip payload when applicable."""
    if _is_gzip_bytes(data):
        try:
            return gzip.decompress(data)
        except Exception as exc:
            logging.getLogger(__name__).warning("Exception suppressed: %s", exc, exc_info=exc)
    return data


def peek_file_source(
    content: bytes | str | os.PathLike,
    filename: str,
) -> tuple[list[str], dict[str, str], int, list[dict]]:
    """Return headers, inferred schema, total row count, and a sample of <=100 rows.

    Accepts either an in-memory ``bytes`` payload or an on-disk path so the
    whole file never has to be loaded at once.  Gzip-compressed payloads are
    decompressed on the fly.
    """
    try:
        from services.file_parser import FileParser
    except ImportError:  # pragma: no cover - compatibility for tests with api root on PYTHONPATH
        from src.services.file_parser import FileParser

    raw_bytes = content if isinstance(content, bytes) else b""
    file_type = FileParser.detect_file_type(filename, raw_bytes or None)

    if file_type in ("csv", "tsv"):
        # Fast path for in-memory payloads that already fit in RAM.
        if isinstance(content, bytes):
            raw = _decompress_bytes_if_gzip(content)
            headers, rows, _enc, _delim = parse_csv_preview(raw, preview_rows=100)
            if not headers:
                raise ValueError("CSV file has no header row")
            total = count_csv_rows(raw)
            sample = [
                dict(zip(headers, (_csv_empty_to_none(c) for c in row)))
                for row in rows[:100]
            ]
            schema = FileParser.infer_schema(sample)
            return headers, schema, total, sample

        # Path-based streaming: read only the preview rows we need and count the
        # rest in a single pass without materializing every cell.
        sample_bytes = _first_bytes(content)
        delim = detect_delimiter(sample_bytes.decode(detect_encoding(sample_bytes), errors="replace"))
        preview_rows: list[list[str]] = []
        total = 0
        headers: list[str] = []
        with _text_reader(content) as reader_file:
            reader = csv.reader(reader_file, delimiter=delim)
            try:
                headers = next(reader)
            except StopIteration:
                raise ValueError("CSV file has no header row") from None
            for row in reader:
                if is_blank_row(row):
                    continue
                total += 1
                if len(preview_rows) < 100:
                    preview_rows.append([_csv_empty_to_none(c) for c in row])
        sample = [dict(zip(headers, row)) for row in preview_rows]
        schema = FileParser.infer_schema(sample)
        return headers, schema, total, sample

    if file_type in ("jsonl", "ndjson"):
        sample_objs: list[dict] = []
        # Ordered union across the *entire* file — keys after sample window must not vanish.
        columns: dict[str, None] = {}
        total = 0
        with _text_reader(content) as reader_file:
            for line in reader_file:
                line = line.strip()
                if not line:
                    continue
                total += 1
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"Invalid JSONL on line {total}: {exc}") from exc
                if not isinstance(obj, dict):
                    raise ValueError(_JSONL_SCALAR_ERROR)
                for k in obj.keys():
                    name = str(k).strip()
                    if name and name not in columns:
                        columns[name] = None
                if len(sample_objs) < 100:
                    sample_objs.append(_json_empty_to_none(obj))
        if total == 0:
            raise ValueError("JSONL file is empty")
        headers = list(columns.keys())
        schema = FileParser.infer_schema(sample_objs)
        return headers, schema, total, sample_objs[:100]

    if file_type == "excel":
        headers, rows, total = _excel_preview(content, preview_rows=100)
        if not headers:
            raise ValueError("Excel file has no header row")
        sample = [dict(zip(headers, row)) for row in rows[:100]]
        schema = FileParser.infer_schema(sample)
        return headers, schema, total, sample

    if file_type == "parquet":
        import pyarrow.parquet as pq
        from services.arrow_schema import schema_from_arrow

        pf = pq.ParquetFile(content) if _is_path(content) else pq.ParquetFile(io.BytesIO(content))
        try:
            total = pf.metadata.num_rows
            headers = [str(c) for c in pf.schema_arrow.names]
            # Prefer writer schema over pandas sample inference (decimals, TZ, nested).
            schema = schema_from_arrow(pf.schema_arrow)
            # to_pylist keeps nested list/struct fidelity — never pandas (ndarray compare lies).
            sample: list[dict] = []
            for batch in pf.iter_batches(batch_size=100):
                sample.extend(batch.to_pylist())
                if len(sample) >= 100:
                    break
            sample = [_json_empty_to_none(r) for r in sample[:100]]
        finally:
            pf.close()
        return headers, schema, total, sample

    if file_type == "orc":
        from pyarrow import orc
        from services.arrow_schema import schema_from_arrow

        if _is_path(content):
            table = orc.ORCFile(content).read()
        else:
            table = orc.read_table(io.BytesIO(content))
        headers = [str(name) for name in table.column_names]
        schema = schema_from_arrow(table.schema)
        total = int(table.num_rows)
        sample = table.slice(0, min(100, total)).to_pylist()
        return headers, schema, total, sample

    if file_type == "avro":
        import fastavro
        from services.avro_schema import schema_map_from_avro

        opener = open(content, "rb") if _is_path(content) else io.BytesIO(content)  # type: ignore[arg-type]
        try:
            reader = fastavro.reader(opener)
            writer_schema = getattr(reader, "writer_schema", None) or getattr(reader, "schema", None)
            schema = schema_map_from_avro(writer_schema) if writer_schema else {}
            sample: list[dict] = []
            total = 0
            for record in reader:
                total += 1
                if not isinstance(record, dict):
                    record = {"value": record}
                if len(sample) < 100:
                    sample.append(record)
                    for k in record.keys():
                        schema.setdefault(str(k), "TEXT")
            headers = list(schema.keys()) if schema else (
                sorted(sample[0].keys()) if sample else []
            )
        finally:
            opener.close()
        return headers, schema, total, sample

    if file_type == "json":
        from services.json_tabular import detect_ijson_records_prefix, load_json_records

        sample_objs: list[dict] = []
        columns: set[str] = set()
        total = 0
        head = _first_bytes(content, 65536)
        prefix = detect_ijson_records_prefix(head)
        if prefix:
            try:
                import ijson
            except ImportError:
                prefix = None
        if prefix:
            with _open_binary(content) as bio:
                for obj in ijson.items(bio, prefix):
                    if not isinstance(obj, dict):
                        continue
                    total += 1
                    columns.update(obj.keys())
                    if len(sample_objs) < 100:
                        sample_objs.append(_json_empty_to_none(obj))
        else:
            raw = content if isinstance(content, (bytes, bytearray)) else None
            if raw is None:
                with _open_binary(content) as bio:
                    raw = bio.read()
            records = load_json_records(bytes(raw))
            total = len(records)
            sample_objs = [_json_empty_to_none(r) for r in records[:100]]
            for obj in records:
                if isinstance(obj, dict):
                    columns.update(obj.keys())
        if total == 0:
            raise ValueError(
                "JSON file has no object rows. Use an array of objects "
                '[{...}], a wrapper like {"data":[{...}]}, or a single object record.'
            )
        headers = sorted(columns)
        schema = FileParser.infer_schema(sample_objs)
        return headers, schema, total, sample_objs[:100]

    raise ValueError(f"File type '{file_type}' does not support streaming ingest")


def _csv_empty_to_none(value: Any) -> Any:
    return None if value == "" else value


def _json_empty_to_none(value: Any) -> Any:
    """JSON/JSONL empty strings are conventionally missing values, not literals."""
    if value == "":
        return None
    if isinstance(value, dict):
        return {k: _json_empty_to_none(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_empty_to_none(v) for v in value]
    return value


def _iter_csv_batches(
    content: bytes | str | os.PathLike,
    chunk_size: int,
):
    sample_bytes = _first_bytes(content)
    enc = detect_encoding(sample_bytes)
    sample = sample_bytes.decode(enc, errors="replace")
    delim = detect_delimiter(sample)
    with _text_reader(content, encoding=enc, newline="") as reader_file:
        reader = csv.DictReader(reader_file, delimiter=delim)
        batch: list[dict] = []
        for row in reader:
            values = dict(row)
            # Must match count_csv_rows, or the source cardinality reconcile
            # compares against is not the set of rows that was read.
            if is_blank_row(values.values()):
                continue
            batch.append({k: _csv_empty_to_none(v) for k, v in values.items()})
            if len(batch) >= chunk_size:
                yield batch
                batch = []
        if batch:
            yield batch


def _iter_jsonl_batches(
    content: bytes | str | os.PathLike,
    chunk_size: int,
):
    with _text_reader(content) as reader_file:
        batch: list[dict] = []
        line_no = 0
        for line in reader_file:
            line = line.strip()
            if not line:
                continue
            line_no += 1
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL on line {line_no}: {exc}") from exc
            if not isinstance(obj, dict):
                raise ValueError(_JSONL_SCALAR_ERROR)
            batch.append(_json_empty_to_none(obj))
            if len(batch) >= chunk_size:
                yield batch
                batch = []
        if batch:
            yield batch


def _iter_json_array_batches(
    content: bytes | str | os.PathLike,
    chunk_size: int,
):
    from services.json_tabular import iter_json_record_dicts

    for batch in iter_json_record_dicts(_open_binary, content, chunk_size=chunk_size):
        yield [_json_empty_to_none(r) for r in batch]


def _batch_iterator_for_type(
    file_type: str,
    content: bytes | str | os.PathLike,
    batch_size: int,
):
    """Return a fresh batch iterator for the given file type.

    Used to re-scan a file from the beginning (e.g. on resume) without mutating
    the primary streaming iterator.  Accepts either ``bytes`` or an on-disk path.
    """
    if file_type in ("csv", "tsv"):
        return _iter_csv_batches(content, batch_size)
    if file_type == "json":
        return _iter_json_array_batches(content, batch_size)
    if file_type == "jsonl" or file_type == "ndjson":
        return _iter_jsonl_batches(content, batch_size)
    if file_type == "excel":
        return _excel_batches(content, batch_size)
    if file_type == "parquet":
        import pyarrow.parquet as pq

        pf = pq.ParquetFile(content) if _is_path(content) else pq.ParquetFile(io.BytesIO(content))

        def _parquet_batches():
            batch: list[dict] = []
            try:
                for record_batch in pf.iter_batches(batch_size=batch_size):
                    # Arrow → pylist preserves nested types; pandas path fails on arrays.
                    for record in record_batch.to_pylist():
                        batch.append(_json_empty_to_none(record))
                        if len(batch) >= batch_size:
                            yield batch
                            batch = []
                if batch:
                    yield batch
            finally:
                pf.close()

        return _parquet_batches()
    if file_type == "orc":
        from pyarrow import orc

        if _is_path(content):
            table = orc.ORCFile(content).read()
        else:
            table = orc.read_table(io.BytesIO(content))

        def _orc_batches():
            batch: list[dict] = []
            for start in range(0, table.num_rows, batch_size):
                chunk = table.slice(start, min(batch_size, table.num_rows - start)).to_pylist()
                batch.extend(chunk)
                while len(batch) >= batch_size:
                    yield batch[:batch_size]
                    batch = batch[batch_size:]
            if batch:
                yield batch

        return _orc_batches()
    if file_type == "avro":
        import fastavro

        def _avro_batches():
            opener = open(content, "rb") if _is_path(content) else io.BytesIO(content)  # type: ignore[arg-type]
            try:
                reader = fastavro.reader(opener)
                batch: list[dict] = []
                for record in reader:
                    if not isinstance(record, dict):
                        record = {"value": record}
                    batch.append(record)
                    if len(batch) >= batch_size:
                        yield batch
                        batch = []
                if batch:
                    yield batch
            finally:
                opener.close()

        return _avro_batches()
    raise ValueError(f"File type '{file_type}' does not support streaming ingest")


def should_stream_file(
    content: bytes | str | os.PathLike,
    filename: str,
    destination: EndpointConfig,
) -> bool:
    if not supports_file_streaming("file", filename, destination):
        return False
    if _is_path(content):
        return True
    if STREAM_THRESHOLD <= 1 and content:
        return True
    if not content:
        return False
    try:
        _, _, total, _ = peek_file_source(content, filename)
        return total >= STREAM_THRESHOLD
    except Exception:
        return False


def stream_file_to_database(
    content: bytes | str | os.PathLike,
    filename: str,
    destination: EndpointConfig,
    mappings: list[dict],
    schema: dict[str, str],
    on_checkpoint: Callable[..., None] | None = None,
    *,
    sync_mode: str = "full_refresh_append",
    stream_contracts: list[dict] | None = None,
    job_id: str | None = None,
    checkpoint: Checkpoint | None = None,
    checkpoint_service: CheckpointService | None = None,
    retry_budget: RetryBudget | None = None,
    backfill_new_fields: bool = False,
    validation_mode: str = "strict",
    source_filter: dict[str, Any] | None = None,
    skip_preflight: bool = False,
    date_locale: str = "",
) -> tuple[int, list[str], dict[str, Any], list[str]]:
    try:
        from services.file_parser import FileParser
    except ImportError:  # pragma: no cover - compatibility for tests with api root on PYTHONPATH
        from src.services.file_parser import FileParser

    file_type = FileParser.detect_file_type(filename)
    columns, probe_schema, total_rows, sample_rows = peek_file_source(content, filename)
    if not schema:
        schema = probe_schema

    # Resolve ambiguous day/month date order from the sample before any transform.
    if not date_locale and sample_rows and columns:
        date_locale = infer_date_locale(sample_rows, columns) or ""
    if date_locale:
        set_active_date_locale(date_locale)

    if not mappings:
        mappings = [{"source": c, "target": c, "confidence": 0.95} for c in columns]

    try:
        from .connector_capabilities import resolve_driver_type
    except ImportError:
        from transfer.connector_capabilities import resolve_driver_type
    dest_type = resolve_driver_type(destination.format)
    dest_cfg = resolve_connector_config(destination)

    from services.value_serializer import json_default

    avg_row_size = 100
    if sample_rows:
        avg_row_size = max(1, int(sum(len(json.dumps(row, default=json_default)) for row in sample_rows) / len(sample_rows)))
    # MongoDB can safely ingest larger batches; keep other destinations under 8 MB
    # to avoid payload limits (e.g. BigQuery streaming insert ~10 MB).
    target_memory_bytes = 64 * 1024 * 1024 if dest_type == "mongodb" else 8 * 1024 * 1024
    batch_size = adaptive_chunk_size(CHUNK_SIZE, avg_row_size, max_size=CHUNK_SIZE, target_memory_bytes=target_memory_bytes)
    # Align file batches to the proxy writer commit size so a dropped socket never
    # straddles tens of thousands of already-committed rows inside one call.
    try:
        from connectors.write_resilience import proxy_stream_batch_size
    except ImportError:
        proxy_stream_batch_size = None  # type: ignore
    if proxy_stream_batch_size is not None:
        batch_size = proxy_stream_batch_size(
            dest_cfg.get("host"),
            connection_string=dest_cfg.get("connection_string")
            or dest_cfg.get("uri")
            or dest_cfg.get("url")
            or "",
            default=batch_size,
        )
    # Object-store writers (S3/GCS/ADLS) emit a single destination object per call.
    # Writing multiple batches would overwrite the same key and silently lose data,
    # so force a single batch — never gate on truthy total_rows (None fails open).
    if dest_type in ("s3", "gcs", "adls"):
        batch_size = max(1, int(total_rows or 10_000_000))
    chunks = max(1, (total_rows + batch_size - 1) // batch_size)
    dest_table = resolve_dest_table(dest_type, destination, "import")

    ddl_log: list[str] = [
        f"STREAM FILE {filename} → {dest_type}.{dest_table} ({total_rows:,} rows, {chunks} batches)",
    ]
    for col in columns:
        ddl_log.append(f"{dest_type.upper()} COLUMN {col} {ddl_type(dest_type, schema.get(col, 'string'))}")

    batch_iter = _batch_iterator_for_type(file_type, content, batch_size)

    column_types = {c: ddl_carrier_type(schema.get(c, "string")) for c in columns}
    target_cols, logical_types = resolve_target_columns(
        mappings, column_types, preserve_case=True
    )
    fingerprint_dest_types = {
        target_cols[i]: logical_types[i] for i in range(len(target_cols))
    }

    try:
        from services.sync_cursor import (
            is_overwrite_sync,
            map_source_to_target,
            requires_upsert,
            resolve_effective_sync_mode,
            resolve_sync_contract,
        )
    except ImportError:
        from src.services.sync_cursor import (
            is_overwrite_sync,
            map_source_to_target,
            requires_upsert,
            resolve_effective_sync_mode,
            resolve_sync_contract,
        )
    contract = resolve_sync_contract(stream_contracts)
    effective_sync = resolve_effective_sync_mode(
        sync_mode,
        contract.sync_mode if contract else None,
    )
    pk_target_cols: list[str] = []
    if contract and contract.primary_key:
        pk_target_cols = [
            map_source_to_target(col, mappings) for col in contract.primary_key_columns()
        ]
    # Object-store destinations (S3/GCS/ADLS) write a single object per call, so
    # row-level upsert keys are not required and the object is overwritten.
    object_store = dest_type in ("s3", "gcs", "adls")
    if object_store and requires_upsert(effective_sync):
        write_mode = "upsert"
    else:
        write_mode = "upsert" if requires_upsert(effective_sync) and pk_target_cols else "insert"
    if requires_upsert(effective_sync) and not pk_target_cols and not object_store:
        raise ValueError(
            f"Sync mode `{effective_sync}` requires primary_key for upsert; "
            "refuse silent insert fallback (set primary_key on the stream contract)"
        )
    # Parallel/chunked resume is only safe with idempotent writes (parity with
    # db stream). Without this, re-processed file chunks append duplicates.
    resuming = bool(checkpoint and getattr(checkpoint, "chunk_index", 0) > 0)
    if resuming and write_mode == "insert" and not object_store:
        if pk_target_cols:
            write_mode = "upsert"
        else:
            raise ValueError(
                "Cannot resume a streaming insert without a primary key. "
                "Re-run this transfer with Full refresh · Overwrite to reload it "
                "safely, or set a primary key to make the replay idempotent."
            )

    checkpoint_service = checkpoint_service or CheckpointService()
    checkpoint = checkpoint or Checkpoint(job_id=job_id or "")
    checkpoint.source_type = "file"
    checkpoint.dest_type = dest_type
    checkpoint.write_mode = write_mode
    checkpoint.conflict_columns = pk_target_cols or []
    checkpoint.chunk_total = chunks
    retry = retry_budget or RetryBudget()
    # File loads are the classic duplicate case: an insert-mode CSV replayed after
    # an ambiguous failure appends the batch a second time. Object-store writes
    # replace the whole object per call, so they are idempotent by construction.
    replay_safety = (
        classify_replay_safety(
            dest_type=dest_type,
            write_mode="replace",
            conflict_columns=["__object__"],
            job_id=job_id,
        )
        if object_store
        else classify_replay_safety(
            dest_type=dest_type,
            write_mode=write_mode,
            conflict_columns=pk_target_cols or None,
            job_id=job_id,
            has_primary_key=bool(pk_target_cols),
        )
    )

    written = checkpoint.rows_processed or 0
    chunk_idx = checkpoint.chunk_index or 0
    resumed = chunk_idx > 0 or written > 0
    dest_summary: dict[str, Any] = {}
    overwrite_keys_acc = begin_overwrite_source_keys(
        effective_sync, pk_target_cols, resumed=resumed
    )
    # Gate-8 append proof needs the cardinality from before the first batch. On a
    # resume the destination already holds rows this job wrote, so the count is
    # no longer a "before" and the delta stays unproven rather than wrong.
    if not resumed:
        rows_before = precount_table(dest_type, dest_cfg, dest_table)
        if rows_before is not None:
            dest_summary[PRECOUNT_KEY] = int(rows_before)
    last_checksum = ""
    # Restore cumulative quarantine counts on resume — Gate-8 conservation is
    # source - (rejected - coerced_null) - skipped, so a resumed pass that starts
    # these at 0 forgets first-pass hold-outs and fails a correct load.
    rejected_total = int(getattr(checkpoint, "rejected_rows", 0) or 0) if resumed else 0
    coerced_null_total = (
        int(getattr(checkpoint, "coerced_null_rows", 0) or 0) if resumed else 0
    )
    # Strict/maximum FAIL-FAST on coercion errors; balanced quarantines them.
    stream_error_policy = transform_error_policy_for_validation_mode(validation_mode)
    warning_samples: list[str] = []
    rejected_details: list[dict] = []

    # Resume: skip chunks that were already committed
    if chunk_idx > 0:
        batch_iter = itertools.islice(batch_iter, chunk_idx, None)

    if source_filter:
        batch_iter = (apply_row_filter(batch, source_filter) for batch in batch_iter)

    fp_accumulator = FingerprintAccumulator()
    # Independent reader cardinality for Gate-8 — never invent from written+held_out.
    source_rows_seen = 0
    batch_quality_enabled = validation_mode in ("strict", "maximum")
    try:
        from services.data_quality import (
            BatchDriftDetector,
            run_integrity_audit,
        )
    except ImportError:  # pragma: no cover - compatibility for tests with api root on PYTHONPATH
        from src.services.data_quality import (
            BatchDriftDetector,
            run_integrity_audit,
        )
    drift_detector = BatchDriftDetector()

    # Phase F6 — align with stream.py (min(4, CPUs); was 2).
    max_workers = int(
        getenv_brand("PARALLEL_WORKERS", str(min(4, os.cpu_count() or 1)))
    )
    # SQLite handles concurrency poorly with a single shared file, so keep it sequential.
    # Snowflake COPY INTO uses a named temporary stage per table; concurrent batches
    # overwrite each other's stage files, so it must also be sequential.
    # Public TCP proxies (Railway, Neon, etc.) drop when multiple bulk writers share
    # the same host — force a single writer connection for those destinations.
    if dest_type in ("sqlite", "snowflake"):
        max_workers = 1
    else:
        try:
            from connectors.write_resilience import is_public_proxy_host
        except ImportError:  # pragma: no cover
            from write_resilience import is_public_proxy_host  # type: ignore
        proxy_host = str(dest_cfg.get("host") or "")
        proxy_cs = str(
            dest_cfg.get("connection_string")
            or dest_cfg.get("uri")
            or dest_cfg.get("url")
            or ""
        )
        if is_public_proxy_host(proxy_host) or is_public_proxy_host(proxy_cs):
            max_workers = 1

    pg_conn_state: dict[str, Any] = {"conn": None}

    def _ensure_pg_conn() -> Any:
        existing = pg_conn_state.get("conn")
        if existing is not None:
            try:
                if getattr(existing, "closed", 0) == 0:
                    return existing
            except Exception:
                pass
            pg_conn_state["conn"] = None
        from connectors.postgresql_conn import get_connection as pg_get_connection

        pg_port = int(
            dest_cfg.get("port")
            or (5439 if dest_type == "redshift" else 5432)
        )
        conn = pg_get_connection(
            host=dest_cfg.get("host", ""),
            port=pg_port,
            database=dest_cfg.get("database", ""),
            username=dest_cfg.get("username", ""),
            password=dest_cfg.get("password", ""),
            connection_string=dest_cfg.get("connection_string", ""),
            ssl=dest_cfg.get("ssl", False),
        )
        try:
            conn.autocommit = False
        except Exception:
            pass
        pg_conn_state["conn"] = conn
        return conn

    from connectors.engine_record_spill import (
        fingerprints_from_spool,
        spill_engine_write_records,
        spool_write_kinds,
    )

    dest_extra = dest_cfg.get("extra") if isinstance(dest_cfg.get("extra"), dict) else {}
    use_source_spool = dest_type in spool_write_kinds()

    def _apply_batch_audit(
        idx: int,
        headers: list[str],
        *,
        rows=None,
        records=None,
    ) -> list[str]:
        local_warnings: list[str] = []
        if not batch_quality_enabled:
            return local_warnings
        audit_sync = "upsert" if write_mode == "upsert" else effective_sync
        audit = run_integrity_audit(
            headers=headers,
            rows=rows,
            records=records,
            column_types=column_types,
            mappings=mappings,
            validation_mode=validation_mode,
            dest_kind=dest_type,
            sync_mode=audit_sync,
        )
        if audit.issues:
            local_warnings.extend(audit.issues[:10])
        if audit.warnings:
            local_warnings.extend(audit.warnings[:10])
        if not audit.passed:
            raise ValueError(
                f"Batch {idx} failed data-quality audit: {'; '.join(audit.issues[:5])}"
            )
        drift_warnings = drift_detector.check(audit.stats or {})
        if drift_warnings:
            if validation_mode == "maximum":
                raise ValueError(
                    f"Batch {idx} drift detected: {'; '.join(drift_warnings[:3])}"
                )
            local_warnings.extend(drift_warnings[:3])
        return local_warnings

    def _spool_fingerprints(spool: Any) -> list[tuple[str, str]]:
        return fingerprints_from_spool(
            spool,
            mappings,
            target_cols,
            headers=list(getattr(spool, "headers", None) or columns or []),
            column_types=column_types,
            dest_db_type=dest_type,
            dest_types=fingerprint_dest_types,
            error_policy=stream_error_policy,
            destination_pk_columns=list(pk_target_cols or []) or None,
            empty_cells_as_null=True,
        )

    def _matrix_fingerprints(
        headers: list[str], data_rows: list[list[Any]]
    ) -> list[tuple[str, str]]:
        mapped_rows, _ = map_rows_for_fingerprint(
            headers=headers,
            data_rows=data_rows,
            mappings=mappings,
            target_cols=target_cols,
            column_types=column_types,
            error_policy=stream_error_policy,
            dest_types=fingerprint_dest_types,
            preserve_case=True,
            dest_kind=dest_type,
            destination_pk_columns=list(pk_target_cols or []) or None,
            empty_cells_as_null=True,
        )
        return (
            row_fingerprints(
                mapped_rows,
                target_cols,
                dest_db_type=dest_type,
                dest_types=fingerprint_dest_types,
            )
            if mapped_rows
            else []
        )

    def _run_file_write(
        idx: int,
        headers: list[str],
        data_rows: list[list[Any]],
        *,
        source_spool: Any = None,
    ) -> tuple[int, str, dict]:
        write_op = partial(
            _write_batch,
            dest_type,
            destination,
            dest_cfg,
            dest_table,
            headers,
            data_rows,
            mappings,
            column_types,
            create_table=(idx == first_index),
            on_checkpoint=None,
            chunk_idx=idx,
            total_chunks=chunks,
            rows_so_far=0,
            write_mode=write_mode,
            conflict_columns=pk_target_cols,
            backfill_new_fields=backfill_new_fields,
            error_policy=stream_error_policy,
            job_id=job_id,
            skip_preflight=skip_preflight,
            # Spreadsheet/CSV blank cells → SQL NULL on nullable typed columns
            # (ITEM 25). DB→DB empty strings still require a Risk Contract.
            empty_cells_as_null=True,
            # Object-store purge vs append-run isolation keys off this. Omitting
            # it left overwrite jobs on the append path so stale part-* objects
            # survived and Gate-8 aggregated mixed generations.
            sync_mode=effective_sync,
            source_spool=source_spool,
            **(
                {
                    "connection": _ensure_pg_conn(),
                    "close_connection": False,
                    "connection_holder": pg_conn_state,
                }
                if dest_type in ("postgresql", "redshift") and max_workers == 1
                else {}
            ),
        )
        try:
            return with_retry(
                write_op,
                budget=RetryBudget(
                    max_attempts=retry.max_attempts,
                    base_delay_seconds=retry.base_delay_seconds,
                    max_delay_seconds=retry.max_delay_seconds,
                    exponential_base=retry.exponential_base,
                    jitter=retry.jitter,
                ),
                replay_safety=replay_safety,
            )
        except WriteBatchBlocked as blocked:
            details = list(blocked.rejected_details or [])
            if details and job_id:
                from services.quarantine_dlq import persist_rejected_rows

                persist_rejected_rows(
                    job_id=str(job_id),
                    rejected_details=details,
                    source="file_stream_batch_abort",
                    connector=str(
                        getattr(destination, "format", None)
                        or getattr(destination, "kind", None)
                        or ""
                    ),
                )
            raise

    def _process_file_chunk(idx: int, batch: list[dict]) -> dict[str, Any]:
        # Worker threads do not inherit the caller's contextvars, so each chunk
        # must re-apply the resolved date locale before any date coercion runs.
        if date_locale:
            set_active_date_locale(date_locale)
        empty = {
            "batch_written": 0,
            "last_checksum": "",
            "dest_summary": {},
            "fingerprints": [],
            "rejected": 0,
            "coerced_null": 0,
            "warnings": [],
            "rejected_details": [],
            "batch_rows": 0,
            "overwrite_keys": [],
        }
        if not batch:
            return empty
        headers = columns or (list(batch[0].keys()) if batch else [])
        # Overwrite keys need the dict rows — collect before spill clears them.
        overwrite_keys = (
            records_to_key_tuples(batch, pk_target_cols, mappings)
            if overwrite_keys_acc is not None
            else None
        )
        if use_source_spool:
            local_warnings = _apply_batch_audit(idx, headers, records=batch)
            spill = spill_engine_write_records(
                batch,
                headers,
                mappings,
                extra=dest_extra,
                clear_records=True,
            )
            try:
                fingerprints = _spool_fingerprints(spill.spool)
                batch_written, last_checksum, dest_summary = _run_file_write(
                    idx, headers, [], source_spool=spill.spool
                )
                batch_rows = spill.unexpanded_row_count
            finally:
                spill.close()
        else:
            headers, data_rows = records_to_matrix(batch, columns)
            local_warnings = _apply_batch_audit(idx, headers, rows=data_rows)
            fingerprints = _matrix_fingerprints(headers, data_rows)
            batch_written, last_checksum, dest_summary = _run_file_write(
                idx, headers, data_rows
            )
            batch_rows = len(data_rows)
        return {
            "batch_written": batch_written,
            "last_checksum": last_checksum,
            "dest_summary": dest_summary,
            "fingerprints": fingerprints,
            "rejected": int(dest_summary.get("rejected_rows", 0) or 0),
            "coerced_null": int(dest_summary.get("coerced_null_rows", 0) or 0),
            "warnings": (dest_summary.get("warnings") or [])[:10] + local_warnings,
            # Full rejected_details for DLQ — never truncate before persist.
            "rejected_details": list(dest_summary.get("rejected_details") or []),
            "batch_rows": batch_rows,
            "overwrite_keys": overwrite_keys,
        }

    first_index = chunk_idx + 1
    batch_enum = enumerate(batch_iter, start=first_index)

    def _apply_file_result(idx: int, result: dict[str, Any]) -> None:
        nonlocal written, rejected_total, coerced_null_total, last_checksum, dest_summary, source_rows_seen
        if overwrite_keys_acc is not None:
            if "overwrite_keys" in result:
                overwrite_keys_acc.observe_tuples(result.get("overwrite_keys"))
            elif result.get("batch_rows"):
                overwrite_keys_acc.observe_tuples(None)
        if result["fingerprints"]:
            fp_accumulator.add_many(result["fingerprints"])
        source_rows_seen += int(result.get("batch_rows") or 0)
        written += result["batch_written"]
        rejected_total += result["rejected"]
        coerced_null_total += result.get("coerced_null", 0)
        warning_samples.extend(result["warnings"])
        new_details = [
            d for d in (result.get("rejected_details") or []) if isinstance(d, dict)
        ]
        rejected_details.extend(new_details)
        last_checksum = result["last_checksum"] or last_checksum
        batch_summary = result.get("dest_summary")
        if isinstance(batch_summary, dict) and batch_summary:
            # Merge batch writer meta; accumulate written_ids across chunks.
            prior_ids = list(dest_summary.get("written_ids") or [])
            # The pre-write count belongs to the FIRST batch: later batches see
            # rows this job already appended, which would hide the delta.
            prior_precount = dest_summary.get(PRECOUNT_KEY)
            dest_summary = dict(batch_summary)
            if prior_precount is not None:
                dest_summary[PRECOUNT_KEY] = prior_precount
            batch_ids = list(batch_summary.get("written_ids") or [])
            if prior_ids or batch_ids:
                merged: list[str] = []
                seen: set[str] = set()
                for x in prior_ids + batch_ids:
                    s = str(x)
                    if not s or s in seen:
                        continue
                    seen.add(s)
                    merged.append(s)
                    if len(merged) >= 500:
                        break
                dest_summary["written_ids"] = merged

        # Persist batch quarantine before continuing — crash must not lose DLQ.
        if new_details and job_id:
            from services.quarantine_dlq import persist_rejected_rows

            persist_rejected_rows(
                job_id=str(job_id),
                rejected_details=new_details,
                source="file_stream_batch",
                connector=str(
                    getattr(destination, "format", None)
                    or getattr(destination, "kind", None)
                    or ""
                ),
            )

        checkpoint.chunk_index = idx
        checkpoint.rows_processed = written
        # Persist cumulative quarantine counts so a resume after a crash restores
        # them and Gate-8 conservation still balances across passes.
        checkpoint.rejected_rows = rejected_total
        checkpoint.coerced_null_rows = coerced_null_total
        checkpoint.checksum = last_checksum
        checkpoint.phase = "writing"
        checkpoint.status = "running"
        # Durable resume requires a real job id. Path / ad-hoc streams with an
        # empty job_id must not hard-fail on job-store reject (no resume contract).
        durable_job = str(job_id or getattr(checkpoint, "job_id", "") or "").strip()
        if durable_job:
            checkpoint.job_id = durable_job
            checkpoint_service.require_save(checkpoint)
        if on_checkpoint:
            on_checkpoint(idx, chunks, written, checkpoint.to_dict())

    try:
        first_idx, first_batch = next(batch_enum)
    except StopIteration:
        raise ValueError("No records found in file")

    try:
        # Process the first batch synchronously so DDL (table/index creation) is
        # committed before any parallel workers try to insert into the new table.
        _apply_file_result(first_idx, _process_file_chunk(first_idx, first_batch))

        with OrderedChunkRunner(max_workers=max_workers) as runner:
            for idx, result in runner.run(batch_enum, _process_file_chunk):
                _apply_file_result(idx, result)
    finally:
        conn = pg_conn_state.get("conn")
        if conn is not None:
            try:
                conn.close()
            except Exception as exc:
                logging.getLogger(__name__).debug(
                    "Exception suppressed: %s", exc, exc_info=exc
                )
            pg_conn_state["conn"] = None

    if written == 0 and rejected_total == 0 and coerced_null_total == 0:
        raise ValueError("No records found in file")
    # All rows may be quarantined (written == 0) — that is a real transfer with
    # DLQ proof, not an empty file. Continue so rejected_details / checksum land.

    # The source checksum has been accumulated incrementally from each batch's
    # mapped fingerprints, so we do not need to parse the entire file a second
    # time.  The FingerprintAccumulator spills to disk above the threshold, so
    # even billion-row transfers stay memory-bounded by a single batch.
    # If the job resumed, we must re-scan the whole file so the fingerprint
    # covers all source rows, not only the ones processed after the checkpoint.
    if resumed and fp_accumulator.total < total_rows:
        full_iter = _batch_iterator_for_type(file_type, content, batch_size)
        # Match the main write path (source_filter applied at read time): count and
        # fingerprint the FILTERED population, or a filtered resume overstates the
        # source count and mis-hashes the checksum against the filtered load.
        if source_filter:
            full_iter = (apply_row_filter(batch, source_filter) for batch in full_iter)
        full_accumulator = FingerprintAccumulator()
        full_source_rows = 0
        for batch in full_iter:
            if not batch:
                continue
            if use_source_spool:
                headers = columns or (list(batch[0].keys()) if batch else [])
                spill = spill_engine_write_records(
                    batch,
                    headers,
                    mappings,
                    extra=dest_extra,
                    clear_records=True,
                )
                try:
                    full_source_rows += spill.unexpanded_row_count
                    fps = _spool_fingerprints(spill.spool)
                    if fps:
                        full_accumulator.add_many(fps)
                finally:
                    spill.close()
            else:
                headers, data_rows = records_to_matrix(batch, columns)
                full_source_rows += len(data_rows)
                fps = _matrix_fingerprints(headers, data_rows)
                if fps:
                    full_accumulator.add_many(fps)
        final_checksum = full_accumulator.digest() if full_accumulator.total else last_checksum
    else:
        full_source_rows = 0
        final_checksum = fp_accumulator.digest() if fp_accumulator.total else last_checksum

    dest_summary["checksum"] = final_checksum or last_checksum
    dest_summary["rejected_rows"] = rejected_total
    dest_summary["coerced_null_rows"] = coerced_null_total
    dest_summary["rejected_details"] = list(rejected_details)
    dest_summary["rejected_details_sample"] = list(rejected_details)[:200]
    dest_summary["warnings"] = warning_samples[:10]
    dest_summary["error_policy"] = "quarantine" if (rejected_total or coerced_null_total) else "none"
    dest_summary["sync_mode"] = effective_sync
    stamp_overwrite_source_keys(dest_summary, overwrite_keys_acc)
    if pk_target_cols:
        dest_summary["conflict_columns"] = list(pk_target_cols)
        dest_summary["primary_key_columns"] = list(pk_target_cols)
    # Stash a bounded source sample so append/upsert Gate-8 reconciliation can
    # perform key-aligned read-back verification instead of failing closed.
    if sample_rows:
        filtered_sample = sample_rows
        if source_filter:
            filtered_sample = apply_row_filter(sample_rows, source_filter)
        dest_summary["reconcile_sample"] = (filtered_sample or [])[:50]
        # Batch PK ids for keyed Gate-8 (full-table digests are not comparable
        # for upsert/append into a non-empty sink).
        if (
            len(pk_target_cols) == 1
            and not dest_summary.get("written_ids")
            and filtered_sample
        ):
            from connectors.writer_common import written_ids_from_mapped_rows

            dest_summary["written_ids"] = written_ids_from_mapped_rows(
                list(filtered_sample),
                list(filtered_sample[0].keys()) if filtered_sample else [],
                pk_target_cols,
            )
    # Reader-side population for Gate-8. Never invent from written + held_out —
    # that circularly balances short reads and hides silent under-delivery. On a
    # resumed run ``source_rows_seen`` counts only the tail after the checkpoint,
    # so the full-file re-scan count (all source rows read) is preferred; using
    # the tail while ``written`` includes prior progress mis-accounts Gate-8.
    if resumed and full_source_rows > 0:
        dest_summary["source_row_count"] = int(full_source_rows)
        dest_summary["source_row_count_source"] = "full_rescan_rows"
    elif source_rows_seen > 0:
        dest_summary["source_row_count"] = int(source_rows_seen)
        dest_summary["source_row_count_source"] = "batch_rows"
    elif int(getattr(fp_accumulator, "total", 0) or 0) > 0:
        dest_summary["source_row_count"] = int(fp_accumulator.total)
        dest_summary["source_row_count_source"] = "fingerprint_accumulator"
    else:
        dest_summary["source_row_count_source"] = "unmeasured"
        dest_summary.pop("source_row_count", None)

    if dest_type in ("postgresql", "mysql", "redshift") and job_id:
        try:
            from connectors.write_resilience import cleanup_write_ledger
        except ImportError:
            cleanup_write_ledger = None  # type: ignore
        if cleanup_write_ledger is not None:
            cleanup_write_ledger(dest_type=dest_type, cfg=dest_cfg, job_id=job_id)

    return written, ddl_log, dest_summary, columns
