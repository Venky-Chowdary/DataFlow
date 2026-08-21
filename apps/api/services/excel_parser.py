"""Excel (.xlsx) parser with streaming row count."""

from __future__ import annotations

import os
from io import BytesIO
from typing import Any, Iterator

from services.read_options import ReadOptions, ReadOptionsError
from services.tabular_rows import is_blank_row
from services.tabular_window import header_and_rows, synthetic_headers
from services.value_serializer import cell_to_string

XLS_UNSUPPORTED_MSG = (
    "Legacy .xls is not supported. Save the workbook as .xlsx and retry."
)

__all__ = [
    "XLS_UNSUPPORTED_MSG",
    "cell_to_string",
    "count_excel_rows",
    "is_blank_row",
    "iter_excel_batches",
    "iter_excel_dicts",
    "list_excel_sheets",
    "parse_excel_preview",
    "require_xlsx",
    "sheet_headers",
    "synthetic_headers",
]


def require_xlsx(path_or_name: str | os.PathLike[str] | bytes | None) -> None:
    """Refuse BIFF .xls — openpyxl only reads Office Open XML (.xlsx)."""
    if path_or_name is None or isinstance(path_or_name, (bytes, bytearray)):
        return
    name = str(path_or_name).lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        raise ValueError(XLS_UNSUPPORTED_MSG)


def sheet_headers(first_row: tuple) -> list[str]:
    """Header names for a sheet's first row, naming unlabelled cells col_N."""
    headers: list[str] = []
    for i, c in enumerate(first_row):
        h = cell_to_string(c).strip() if c is not None else ""
        headers.append(h if h else f"col_{i}")
    return headers


def _select_sheet(wb: Any, options: ReadOptions) -> Any:
    """The worksheet the options name, or a refusal that lists the real names.

    Silently falling back to the active sheet would transfer the wrong data
    under the right job name, which is worse than not transferring at all.
    """
    names = list(getattr(wb, "sheetnames", []) or [])
    if options.sheet:
        if options.sheet in names:
            return wb[options.sheet]
        folded = {str(n).strip().casefold(): n for n in names}
        match = folded.get(options.sheet.strip().casefold())
        if match is not None:
            return wb[match]
        available = ", ".join(f"'{n}'" for n in names) or "none"
        raise ReadOptionsError(
            f"Workbook has no sheet named '{options.sheet}'. Available: {available}"
        )
    if options.sheet_index >= 0:
        if options.sheet_index >= len(names):
            raise ReadOptionsError(
                f"Workbook has {len(names)} sheet(s); sheet_index "
                f"{options.sheet_index} is out of range. Available: "
                + (", ".join(f"[{i}] '{n}'" for i, n in enumerate(names)) or "none")
            )
        return wb[names[options.sheet_index]]
    return wb.active


def _sheet_header_and_rows(
    ws: Any, options: ReadOptions
) -> tuple[list[str], Iterator[Any]]:
    """Header names plus the sheet's data rows inside the declared window."""
    return header_and_rows(
        ws.iter_rows(values_only=True),
        options,
        header_names=sheet_headers,
        source_label=f"Sheet '{getattr(ws, 'title', '')}'",
    )


def list_excel_sheets(content: bytes | Any) -> list[dict[str, Any]]:
    """Sheet inventory for the source picker: name, position, and a first-row peek.

    ``max_row``/``max_column`` are the used range and formatting inflates them,
    so they are reported as ``used_rows``/``used_columns`` and never as a row
    count. The row count is a COUNT, and a COUNT scans.
    """
    wb = _load_workbook(content)
    try:
        active = getattr(getattr(wb, "active", None), "title", "")
        sheets: list[dict[str, Any]] = []
        for index, name in enumerate(list(getattr(wb, "sheetnames", []) or [])):
            ws = wb[name]
            first = next(ws.iter_rows(values_only=True), None) or ()
            sheets.append(
                {
                    "name": name,
                    "index": index,
                    "is_active": name == active,
                    "used_rows": int(getattr(ws, "max_row", 0) or 0),
                    "used_columns": int(getattr(ws, "max_column", 0) or 0),
                    "first_row": [cell_to_string(c) for c in first],
                }
            )
        return sheets
    finally:
        wb.close()


def _load_workbook(content: bytes | Any):
    """openpyxl workbook from bytes or a seekable binary handle.

    Dest gzip Excel spools a decompressed image (workbook formats are not
    sequential). ``load_workbook`` already accepts a file-like; wrapping
    that image in a second ``BytesIO`` would be a third copy.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Excel import is not ready on this platform node. Datawrap bundles file parsers — retry shortly."
        ) from exc
    if isinstance(content, (bytes, bytearray)):
        stream: Any = BytesIO(content)
    elif isinstance(content, (str, os.PathLike)):
        # An on-disk image streams from the filesystem; wrapping it in bytes
        # first would defeat the spill that put it on disk.
        require_xlsx(content)
        stream = content
    else:
        stream = content
        try:
            stream.seek(0)
        except Exception as exc:
            raise ValueError("Excel workbook source is not seekable") from exc
    return load_workbook(stream, read_only=True, data_only=True)


def parse_excel_preview(
    content: bytes,
    preview_rows: int = 100,
    options: ReadOptions | None = None,
) -> tuple[list[str], list[list[str]], int]:
    opts = options or ReadOptions()
    wb = _load_workbook(content)
    try:
        ws = _select_sheet(wb, opts)
        if ws is None:
            return [], [], 0

        headers, rows = _sheet_header_and_rows(ws, opts)
        if not headers:
            return [], [], 0

        preview: list[list[str]] = []
        total = 0
        for row in rows:
            total += 1
            if len(preview) < preview_rows:
                preview.append([cell_to_string(c) for c in row])
        return headers, preview, total
    finally:
        wb.close()


def iter_excel_dicts(content: bytes | Any, options: ReadOptions | None = None) -> Iterator[dict]:
    """Value-bearing Excel records. Same population as ``count_excel_rows``.

    Header is not a record. ``is_blank_row`` (formatting-only used-range)
    is not a record. Extra cells beyond the header refuse silent column
    drop — ingest already raised; Gate-8 must not hash a truncated row.
    ``max_row`` is not dest population. Gate-8 cell checksum and dest
    sample walk this iterator. ``options`` narrows the window (sheet, header
    row, head/tail skips) and every caller must pass the same one, or the
    population Validate profiled is not the population the writer sends.
    """
    opts = options or ReadOptions()
    wb = _load_workbook(content)
    ws = _select_sheet(wb, opts)
    if ws is None:
        wb.close()
        return
    try:
        headers, rows = _sheet_header_and_rows(ws, opts)
        if not headers:
            return
        for row in rows:
            if len(row) > len(headers):
                raise ValueError(
                    f"Excel row has {len(row)} cells but header has {len(headers)} "
                    "columns; refuse silent column drop — widen the header row "
                    "or fix the sheet"
                )
            yield {
                headers[i]: cell_to_string(c)
                for i, c in enumerate(row[: len(headers)])
            }
    finally:
        wb.close()


def iter_excel_batches(
    content: bytes, chunk_size: int, options: ReadOptions | None = None
) -> Iterator[list[dict]]:
    """Stream Excel rows as dict batches without loading the full sheet into RAM."""
    batch: list[dict] = []
    for record in iter_excel_dicts(content, options):
        batch.append(record)
        if len(batch) >= chunk_size:
            yield batch
            batch = []
    if batch:
        yield batch


def count_excel_rows(content: bytes | Any, options: ReadOptions | None = None) -> int:
    """Count rows that carry values.

    ``max_row`` is the used range, which formatting inflates; using it as the
    source cardinality makes reconciliation compare against rows that were
    never read. Extra cells beyond the header raise — dest COUNT then
    stays unmeasured rather than hashing a truncated row.
    """
    return sum(1 for _ in iter_excel_dicts(content, options))
